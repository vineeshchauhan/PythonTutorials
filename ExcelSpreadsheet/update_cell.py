import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']

#cell
cell = sheet['a1']

cell1 = sheet.cell(1,1)

print(cell.value)
print(cell1.value)


def update_cell():
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * .9
        sheet.cell(row, 4).value = corrected_price


update_cell()

wb.save('transactions1.xlsx')